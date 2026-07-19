#!/usr/bin/env python3
"""Generate video-ready visuals from a deals file (CSV or XLSX).

Produces a set of PNG cards/charts sized for short-form video, ready to drop
into a TikTok / Reels / Shorts edit (or the landscape variant for YouTube):

  01_headline.png      Total raised, deal count, growth
  02_stage_split.png   Dollars by stage (bar)
  03_regions.png       Top regions by dollars (bar)
  04_hiring_gtm.png    Series A-C companies hiring Sales & Marketing (list card)

Usage:
  python scripts/make_visuals.py <deals.csv|deals.xlsx> \
      [--prior-total USD] [--format vertical|landscape|square] [--out outputs/visuals]

Requires: matplotlib. Reads the same schema as assets/deals-template.csv.
"""
import argparse
import csv
import os
import sys

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib import font_manager  # noqa: F401

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from aggregate import aggregate, fmt_usd  # reuse the metrics logic

# --- palette (calm, friendly, high-contrast for video) ---
BG = "#0F1626"
CARD = "#0F1626"
INK = "#F5F7FA"
MUTE = "#9AA5B8"
ACCENT = "#5B8DEF"
ACCENT2 = "#57C7A3"
UP = "#57C7A3"
DOWN = "#EF6F6C"

SIZES = {           # width, height in inches @ 100 dpi -> pixels
    "vertical": (10.8, 19.2),   # 1080x1920
    "landscape": (19.2, 10.8),  # 1920x1080
    "square": (10.8, 10.8),     # 1080x1080
}


def load_rows(path):
    if path.lower().endswith((".xlsx", ".xlsm")):
        from openpyxl import load_workbook
        wb = load_workbook(path, data_only=True)
        ws = wb["Deals"] if "Deals" in wb.sheetnames else wb.active
        rows = list(ws.iter_rows(values_only=True))
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        out = []
        for r in rows[1:]:
            if all(c is None for c in r):
                continue
            out.append({headers[i]: ("" if v is None else v) for i, v in enumerate(r) if i < len(headers)})
        return out
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def new_fig(fmt):
    w, h = SIZES[fmt]
    fig = plt.figure(figsize=(w, h), dpi=100)
    fig.patch.set_facecolor(BG)
    ax = fig.add_axes([0, 0, 1, 1])
    ax.set_facecolor(CARD)
    ax.axis("off")
    return fig, ax


def footer(ax, label):
    ax.text(0.06, 0.03, label, transform=ax.transAxes, color=MUTE,
            fontsize=16, ha="left", va="bottom")


def save(fig, out, name):
    path = os.path.join(out, name)
    fig.savefig(path, facecolor=fig.get_facecolor())
    plt.close(fig)
    return path


def card_headline(m, out, fmt, tag):
    fig, ax = new_fig(fmt)
    ax.text(0.06, 0.90, "THIS PERIOD IN VENTURE", color=ACCENT, fontsize=26,
            fontweight="bold", transform=ax.transAxes)
    ax.text(0.06, 0.80, fmt_usd(m["total_raised_usd"]), color=INK, fontsize=96,
            fontweight="bold", transform=ax.transAxes)
    ax.text(0.06, 0.73, "raised across " + str(m["funding_deal_count"]) + " disclosed rounds",
            color=MUTE, fontsize=28, transform=ax.transAxes)
    g = m["growth_pct_vs_prior"]
    if g is not None:
        col = UP if g >= 0 else DOWN
        arrow = "\u25B2" if g >= 0 else "\u25BC"
        ax.text(0.06, 0.60, f"{arrow} {abs(g):.0f}% vs prior period", color=col,
                fontsize=34, fontweight="bold", transform=ax.transAxes)
    if m["acquisition_count"]:
        ax.text(0.06, 0.52, f"+ {m['acquisition_count']} acquisitions", color=MUTE,
                fontsize=26, transform=ax.transAxes)
    footer(ax, tag)
    return save(fig, out, "01_headline.png")


def bar_card(title, pairs, out, fmt, name, tag, value_fmt=fmt_usd, color=ACCENT):
    pairs = [(k, v) for k, v in pairs if v]
    fig, ax0 = new_fig(fmt)
    ax0.text(0.06, 0.92, title, color=ACCENT, fontsize=30, fontweight="bold",
             transform=ax0.transAxes)
    if not pairs:
        ax0.text(0.06, 0.5, "No data", color=MUTE, fontsize=28, transform=ax0.transAxes)
        footer(ax0, tag)
        return save(fig, out, name)
    ax = fig.add_axes([0.30, 0.12, 0.63, 0.72])
    ax.set_facecolor(CARD)
    labels = [k for k, _ in pairs][::-1]
    vals = [v for _, v in pairs][::-1]
    bars = ax.barh(labels, vals, color=color, height=0.62)
    ax.set_xlim(0, max(vals) * 1.18)
    for sp in ax.spines.values():
        sp.set_visible(False)
    ax.tick_params(axis="y", colors=INK, labelsize=26, length=0)
    ax.tick_params(axis="x", colors=BG, length=0)
    ax.set_xticks([])
    for b, v in zip(bars, vals):
        ax.text(b.get_width() + max(vals) * 0.02, b.get_y() + b.get_height() / 2,
                value_fmt(v), color=INK, fontsize=24, va="center", fontweight="bold")
    footer(ax0, tag)
    return save(fig, out, name)


def card_hiring(hiring, out, fmt, tag):
    fig, ax = new_fig(fmt)
    ax.text(0.06, 0.93, "WHO'S HIRING GTM", color=ACCENT, fontsize=30,
            fontweight="bold", transform=ax.transAxes)
    ax.text(0.06, 0.885, "Series A\u2013C companies staffing up Sales & Marketing",
            color=MUTE, fontsize=22, transform=ax.transAxes)
    if not hiring:
        ax.text(0.06, 0.5, "None captured this period", color=MUTE, fontsize=26,
                transform=ax.transAxes)
        footer(ax, tag)
        return save(fig, out, "04_hiring_gtm.png")
    y = 0.80
    step = min(0.11, 0.66 / max(len(hiring), 1))
    for h in hiring[:7]:
        ax.text(0.06, y, h["company"], color=INK, fontsize=34, fontweight="bold",
                transform=ax.transAxes)
        ax.text(0.06, y - step * 0.42, f"{h['stage']}  \u00b7  hiring {h['roles']}",
                color=ACCENT2, fontsize=22, transform=ax.transAxes)
        y -= step
    footer(ax, tag)
    return save(fig, out, "04_hiring_gtm.png")


def main():
    p = argparse.ArgumentParser()
    p.add_argument("deals_path")
    p.add_argument("--prior-total", type=float, default=None)
    p.add_argument("--format", choices=list(SIZES), default="vertical")
    p.add_argument("--out", default="outputs/visuals")
    p.add_argument("--tag", default="Funding Digest", help="Footer label, e.g. the date window")
    args = p.parse_args()

    os.makedirs(args.out, exist_ok=True)
    rows = load_rows(args.deals_path)
    m = aggregate(rows, prior_total=args.prior_total)

    made = []
    made.append(card_headline(m, args.out, args.format, args.tag))
    made.append(bar_card("FUNDING BY STAGE", list(m["stage_dollars"].items()),
                         args.out, args.format, "02_stage_split.png", args.tag))
    made.append(bar_card("TOP REGIONS BY $", list(m["dollars_by_region"].items()),
                         args.out, args.format, "03_regions.png", args.tag, color=ACCENT2))
    made.append(card_hiring(m["series_abc_hiring_gtm"], args.out, args.format, args.tag))

    print("Generated:")
    for pth in made:
        print(" ", pth)


if __name__ == "__main__":
    main()
